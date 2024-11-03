from flask import Flask, render_template, request, jsonify, redirect, url_for
import time
import pr

import openpyxl
import openpyxl.workbook

filename = 'Main.xlsx'
try:
   workbook = openpyxl.load_workbook(filename)
   sheet = workbook.active
   
except Exception:
   workbook = openpyxl.Workbook()
   sheet = workbook.active
   workbook.save(filename)

if sheet['A1'] == 'User_Ty':
    workbook.save(filename)

else:
    sheet['A1'] = 'User_Ty'
    sheet['B1'] = 'Pass_Ty'
    sheet['C1'] = 'User'
    workbook.save(filename)




if sheet.max_row == 1:
   m_rows = 1

else:
   m_rows = sheet.max_row+1

user_ty = []
pass_ty = []
user = []



def fill_column(filename, column_letter, data):


    #workbook = openpyxl.load_workbook(filename)
    #sheet = workbook.active

    for i, value in enumerate(data):
        cell = sheet[column_letter + str((m_rows) + i)]
        cell.value = value
        sheet[f'C{sheet.max_row}'].value =  user[-1]

    workbook.save(filename)

#res_factor = time.time()

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process_text', methods=['POST'])
def process_text():
    start_time = time.time()
    #time.sleep(5)
    text = request.json['text']
    user.append(str(text))
    


    #fill_column(filename, col_letter)
    #p = len(text) + 1
    #password = request.json['text']
    
    try:
        end_time = time.time()  + len(text) - 2
        typing_speed = round(len(text) / (end_time - start_time), 4)
        if typing_speed > 0 and typing_speed != 3.0:
           user_ty.append(float(typing_speed))

    except Exception:
        end_time = time.time() + 5
        typing_speed = round(len(text) / (end_time - start_time), 4)
        if typing_speed > 0 and typing_speed != 3.0:
          user_ty.append(float(typing_speed))


    
    #user_ty.append(typing_speed)
    end_time = start_time = 0


    #print(numpy.random.choice(user_ty))
    
    

    return jsonify({'typing_speed': typing_speed})

@app.route('/progress')
def pro1():
    global user_tyc
    global pass_tyc
    user_tyc = []
    
    pass_tyc = []
    
    user_tyc.append(float(user_ty[-1]))
    pass_tyc.append(float(pass_ty[-1]))

    col_letter1 = 'A'
    col_letter2 = 'B'
    col_letter3 = 'C'

    fill_column(filename, col_letter1, user_tyc)
    fill_column(filename, col_letter2, pass_tyc)
    #fill_column(filename, col_letter3, user)
    print(user[-1])
    
    i = True

    if i :
        return redirect(url_for('pro'))
        



   

    
    #print(len(user_ty), len(pass_ty))
    #workbook.save(filename)

    return render_template('index3.html')


@app.route('/pro')
def pro():
    
    intake = [[float(user_tyc[0]), float(pass_tyc[0])]]
    pr.process(epochs=20)
    xr = pr.resultant(intake=intake)

    if xr == 1:
        xr = 'Nandhan'

    else:
        #xr == 0:
        xr = 'Person1'

    data = {'user' : xr}

    return render_template('index3.html', data=data)



@app.route('/process')
def index2():

    return render_template('index2.html')


@app.route('/process_pass', methods=['POST'])
def main():
    start_time = time.time()

    text = request.json['text']
    try:
        end_time = time.time()  + len(text) - 2
        typing_speed = round(len(text) / (end_time - start_time), 4)
        if typing_speed > 0 and typing_speed != 3.0:
           pass_ty.append(float(typing_speed))

    except Exception:
        end_time = time.time() + 5
        typing_speed = round(len(text) / (end_time - start_time), 4)
        if typing_speed > 0 and typing_speed != 3.0:
          pass_ty.append(float(typing_speed))
    

    end_time = start_time = 0

    return jsonify({'typing_speed' : typing_speed})

if __name__ == '__main__':
    app.run(debug=True)